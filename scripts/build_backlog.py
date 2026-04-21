"""
Generates BACKLOG.xlsx — project task tracker.
Re-run this script to regenerate from source-of-truth list below.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

TASKS = [
    # (id, area, priority, effort, title, description, status)
    ("B-01", "Portfolio", "High", "L",
     "Rewrite project pages as case studies",
     "SCAFFOLDED + placeholder content filled for all 9 projects EN+RU. Pages live at /works/[slug] and /ru/works/[slug] with full case study layout. View-transition morph from gallery cover wired. NOW NEEDS: replace placeholder facts/quotes/metrics with real client data — see B-19 for per-project content pass.",
     "In progress"),

    ("B-19", "Content", "High", "L",
     "Replace placeholder case study content with real data",
     "All 9 case studies have filled placeholder content (src/content/works/{en,ru}/*.md). Need real pass: verify years/roles/stacks, add real client quotes (testimonials), add real outcome metrics, add gallery images per project, confirm or add live URLs. Do one project at a time — change draft: false to draft: true to hide unfinished ones if needed.",
     "Backlog"),

    ("B-02", "Performance", "Medium", "L",
     "Migrate Spline scene to raw three.js",
     "Spline viewer ships ~2-3MB and hurts first paint on mobile/4G. Export the elk scene and render with three.js directly. Keeps the visual, cuts the weight.",
     "Backlog"),

    ("B-03", "SEO/Social", "Medium", "M",
     "OG-image + full social meta",
     "Generate an OG preview image (1200x630) with 3D elk + Elkhovering wordmark. Add og:title, og:description, og:image, og:url, twitter:card meta tags in Layout. Affects how links look in Telegram, Twitter, LinkedIn.",
     "Backlog"),

    ("B-04", "Copy", "High", "S",
     "Finalize hero tagline per new positioning",
     "DONE. Hero H1 -> 'Creative director', tagline -> 'The art director you hire before you need a team'. Applied in en.json, ru.json, Hero.astro fallback.",
     "Done"),

    ("B-05", "Copy", "High", "M",
     "Rewrite About as manifesto (reuse resume copy)",
     "DONE. About rewritten using CV copy: intersection of IT/product/digital, 300+ projects, walk into chaos, leave stable platform. Personality bit now: economic strategy games + gardening + game design at Syrnik.dev.",
     "Done"),

    ("B-06", "Services", "Medium", "M",
     "Reduce services 6 -> 2-3",
     "Current list (creative sites / brand / NFT / consulting / team lead / lectures) spans 3 different businesses. With positioning now fixed ('art director before the team'), consolidate into primary offer (fractional digital lead) + 1-2 adjacent.",
     "Backlog"),

    ("B-07", "Trust", "High", "M",
     "Add testimonials section",
     "Pull 2-3 short quotes from the 3 recommendations in CV (Мастер рекламы, Р-Групп, Nut Valley Armenia). Add a clean testimonials block between Works and Services or between Services and Contacts.",
     "Backlog"),

    ("B-08", "Trust", "Medium", "S",
     "Add client logos row",
     "Even 5-6 monochrome logos of past clients (Heavens.pro alumni, Syrnik, Custom, etc) gives instant credibility. Usually sits above-the-fold or in About.",
     "Backlog"),

    ("B-09", "CTA", "Medium", "S",
     "Resolve LinkedIn / Behance empty links",
     "Social row in Contacts still has empty href=\"\" for linkedin and behance. Either remove them, or point to real profiles. Currently dead clicks.",
     "Backlog"),

    ("B-17", "CMS", "High", "L",
     "Admin panel for editing case studies",
     "Content lives in src/content/works/{en,ru}/*.md with strict Zod schema — editable in any git-based CMS. Research & pick: Decap CMS (open, git-based, needs GitHub OAuth proxy for shared hosting), Tina (visual, free tier), Sanity/Contentful (external API). Must preserve existing frontmatter fields, support EN+RU parallel editing, and let non-devs add galleries/update URLs/toggle draft without touching git directly.",
     "Backlog"),

    ("B-18", "i18n", "Medium", "M",
     "Unify i18n strategy (client-side vs URL-based)",
     "DONE. Unified on client-side i18n via html[data-lang] across the whole site. Both EN and RU blocks coexist in every HTML page (chrome via data-i18n, case studies via [data-case-lang] wrappers), CSS toggles visibility. Inline early script in <head> sets data-lang from localStorage before first paint to avoid FOUC. Old /ru/ paths redirected for backwards compat.",
     "Done"),

    ("B-16", "CTA", "High", "M",
     "Redesign Contacts — Telegram-primary without breaking composition",
     "Current Contacts restored to original (big H1 = copyable email, small row = or/telegram/linkedin/behance). But Telegram is the actual primary CTA per positioning. Need a redesigned block that makes Telegram the hero action while preserving the designer-grade composition Eugene values. Likely variations to prototype: (a) big @elkhovering H1 + email as secondary copyable, (b) split H1 with both, (c) asymmetric two-column. Design-first, not markup-first.",
     "Backlog"),

    ("B-10", "Deploy", "High", "M",
     "Update GitHub Actions deploy for Astro build",
     "DONE. Workflow now runs npm ci + npm run build, then SFTPs dist/ to public_html with mirror semantics — uploads build output and prunes orphans (old index.html, style.css, script.js from pre-Astro). Preserves cgi-bin / .well-known / server-managed .htaccess.",
     "Done"),

    ("B-11", "Cleanup", "Low", "S",
     "Remove legacy root files after Astro is verified",
     "Delete index.html, style.css, mobile.css, script.js, nav.html, works.html from repo root once the Astro build is confirmed working in production. Keep as git history.",
     "Backlog"),

    ("B-12", "Features", "Low", "M",
     "Contact form as alternative to Telegram",
     "For clients who prefer email-first flow. Use Formspree or similar (no backend on shared hosting). Low priority since Telegram is primary CTA.",
     "Backlog"),

    ("B-13", "Content", "Medium", "L",
     "Per-project data: role, stack, year, outcome",
     "DONE. Schema finalized in src/content.config.ts — frontmatter covers slug/title/summary/year/role/url/client/stack/tags/cover/gallery/draft. Layout-only meta (col/row/rellax/layer) stays in src/data/works.js. Filling real values is a content task under B-01.",
     "Done"),

    ("B-14", "UX", "Low", "S",
     "Mobile layout pass",
     "Current mobile.css handles basics but wasn't tested against new Astro components. Do a proper responsive review when content is final.",
     "Backlog"),

    ("B-15", "Analytics", "Low", "S",
     "Add lightweight analytics",
     "Plausible or Umami (privacy-first, no cookie banner needed). Just to know which works get clicks and where visitors come from. Not GA.",
     "Backlog"),
]

AREA_COLORS = {
    "Portfolio":   "FFE5CC",
    "Performance": "FFCCCC",
    "SEO/Social":  "E5CCFF",
    "Copy":        "CCE5FF",
    "Services":    "CCE5FF",
    "Trust":       "CCFFCC",
    "CTA":         "FFFFCC",
    "Deploy":      "E5E5E5",
    "Cleanup":     "E5E5E5",
    "Features":    "FFE5F2",
    "Content":     "FFE5CC",
    "UX":          "CCFFFF",
    "Analytics":   "E5E5E5",
    "CMS":         "E5CCFF",
    "i18n":        "CCE5FF",
}
PRIORITY_COLORS = {
    "High":   "FFCCCC",
    "Medium": "FFF2CC",
    "Low":    "E5E5E5",
}
STATUS_COLORS = {
    "Backlog":     "FFFFFF",
    "In review":   "FFF2CC",
    "In progress": "CCE5FF",
    "Blocked by B-01": "E5E5E5",
    "Done":        "CCFFCC",
}

def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Backlog"

    headers = ["ID", "Area", "Priority", "Effort", "Task", "Description", "Status"]
    ws.append(headers)

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="222222", end_color="222222", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF", size=11)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font_white
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center")

    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for t in TASKS:
        ws.append(t)

    for row_idx in range(2, len(TASKS) + 2):
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = border
        area = ws.cell(row=row_idx, column=2).value
        prio = ws.cell(row=row_idx, column=3).value
        status = ws.cell(row=row_idx, column=7).value
        if area in AREA_COLORS:
            ws.cell(row=row_idx, column=2).fill = PatternFill(start_color=AREA_COLORS[area], end_color=AREA_COLORS[area], fill_type="solid")
        if prio in PRIORITY_COLORS:
            ws.cell(row=row_idx, column=3).fill = PatternFill(start_color=PRIORITY_COLORS[prio], end_color=PRIORITY_COLORS[prio], fill_type="solid")
        if status in STATUS_COLORS:
            ws.cell(row=row_idx, column=7).fill = PatternFill(start_color=STATUS_COLORS[status], end_color=STATUS_COLORS[status], fill_type="solid")

    widths = [6, 14, 10, 8, 42, 90, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 22
    for r in range(2, len(TASKS) + 2):
        ws.row_dimensions[r].height = 60

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save("BACKLOG.xlsx")
    print(f"Wrote BACKLOG.xlsx with {len(TASKS)} tasks")

if __name__ == "__main__":
    build()
